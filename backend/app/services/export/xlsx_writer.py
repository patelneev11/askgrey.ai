from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from app.services.pdf_extraction import CellStatus, ExtractionTable

from .layout import (
    SOURCES_HEADERS,
    CitationEntry,
    cell_of,
    cell_text,
    citation_entries,
    clean,
    headers,
    paper_name,
    refs_by_cell,
    validate,
)
from .models import XLSX_MEDIA_TYPE, ExportFile, ExportOptions

DATA_SHEET = "Review table"
SOURCES_SHEET = "Sources"

HEADER_FILL = PatternFill("solid", fgColor="14181F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="1F6FEB", underline="single")
UNVERIFIED_FONT = Font(color="9A6700", italic=True)

DEFAULT_WIDTH = 28.0
WIDTHS: dict[str, float] = {"Paper": 46, "Source": 34, "Pages": 8, "Row status": 13}
SOURCES_WIDTHS: dict[str, float] = {"Ref": 7, "Quote": 70, "Page": 6, "Match": 12, "Value": 26}


def _text_cell(sheet: Worksheet, row: int, column: int, value: str) -> Cell:
    """
    Write a string, pinning the cell to the text type.

    openpyxl infers the type from the value, and a string beginning with `=` is stored as a
    live formula. An extracted value is untrusted text out of a third-party PDF, so the type
    is pinned rather than the text being mangled with an escape prefix as the CSV writer has
    to do.
    """
    cell = sheet.cell(row=row, column=column)
    cell.value = clean(value)
    cell.data_type = "s"
    return cell


def _write_header(sheet: Worksheet, labels: list[str], widths: dict[str, float]) -> None:
    for index, label in enumerate(labels, start=1):
        cell = _text_cell(sheet, 1, index, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(label, DEFAULT_WIDTH)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(labels))}1"
    sheet.row_dimensions[1].height = 22


def _write_sources(sheet: Worksheet, entries: list[CitationEntry]) -> None:
    _write_header(sheet, list(SOURCES_HEADERS), SOURCES_WIDTHS)
    for offset, entry in enumerate(entries):
        row = offset + 2
        for column, value in enumerate(entry.as_row(), start=1):
            if isinstance(value, int):
                sheet.cell(row=row, column=column).value = value
            else:
                _text_cell(sheet, row, column, value)
        sheet.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
        if entry.source_url:
            link = sheet.cell(row=row, column=len(SOURCES_HEADERS))
            link.hyperlink = entry.source_url
            link.font = LINK_FONT


def _tooltip(entry: CitationEntry) -> str:
    """Hover text on a cited cell: the quote, without leaving the data sheet."""
    quote = entry.quote if len(entry.quote) <= 250 else entry.quote[:249] + "…"
    suffix = "" if entry.match == "exact" else f" ({entry.match} match)"
    return f"{entry.ref} · p{entry.page}{suffix}\n{quote}"


def write_xlsx(table: ExtractionTable, options: ExportOptions | None = None) -> ExportFile:
    """
    Render a review table as a workbook.

    Citations live on their own `Sources` sheet, one row per cited cell, and every cited
    value on the data sheet is an internal hyperlink into that row (with the quote as its
    hover tooltip). See the module README for why this beats per-cell comments.
    """
    options = options or ExportOptions()
    validate(table, options)

    workbook = Workbook()
    data = workbook.active
    if data is None:  # pragma: no cover - openpyxl always creates one sheet
        data = workbook.create_sheet()
    data.title = DATA_SHEET

    labels = headers(table, options)
    _write_header(data, labels, WIDTHS)
    leading = len(labels) - len(table.columns)

    entries = citation_entries(table) if options.include_citations else []
    by_cell = refs_by_cell(entries)
    sources = workbook.create_sheet(SOURCES_SHEET) if options.include_citations else None
    rows_by_ref = {entry.ref: index + 2 for index, entry in enumerate(entries)}

    for row_index, row in enumerate(table.rows, start=1):
        excel_row = row_index + 1
        _text_cell(data, excel_row, 1, paper_name(row))
        if options.include_metadata:
            _text_cell(data, excel_row, 2, row.source_url)
            data.cell(row=excel_row, column=3).value = row.page_count
            _text_cell(data, excel_row, 4, row.status.value)
            if row.source_url:
                link = data.cell(row=excel_row, column=2)
                link.hyperlink = row.source_url
                link.font = LINK_FONT

        for offset, column in enumerate(table.columns):
            excel_column = leading + offset + 1
            cell = cell_of(row, column.key)
            _text_cell(data, excel_row, excel_column, cell_text(cell))
            written = data.cell(row=excel_row, column=excel_column)
            if cell.status is CellStatus.UNGROUNDED:
                written.font = UNVERIFIED_FONT
                continue
            entry = by_cell.get((row_index, column.key))
            if entry is None:
                continue
            written.hyperlink = Hyperlink(
                ref=written.coordinate,
                location=f"'{SOURCES_SHEET}'!A{rows_by_ref[entry.ref]}",
                tooltip=_tooltip(entry),
            )
            written.font = LINK_FONT

    if sources is not None:
        _write_sources(sources, entries)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return ExportFile(
        filename=f"{options.filename_stem}.xlsx",
        media_type=XLSX_MEDIA_TYPE,
        content=buffer.getvalue(),
    )
