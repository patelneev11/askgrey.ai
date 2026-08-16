from __future__ import annotations

import io
import time

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.export import (
    DATA_SHEET,
    SOURCES_SHEET,
    XLSX_MEDIA_TYPE,
    ExportOptions,
    TableTooLargeError,
    write_xlsx,
)
from app.services.export.models import MAX_COLUMNS
from app.services.pdf_extraction import ExtractionField, MatchQuality
from tests.export.conftest import grounded, paper, sample_table, table, ungrounded


def sheets(content: bytes) -> tuple[Worksheet, Worksheet]:
    workbook = load_workbook(io.BytesIO(content))
    return workbook[DATA_SHEET], workbook[SOURCES_SHEET]


def test_workbook_has_a_data_sheet_and_a_sources_sheet() -> None:
    rendered = write_xlsx(sample_table())

    assert rendered.filename == "review-table.xlsx"
    assert rendered.media_type == XLSX_MEDIA_TYPE
    workbook = load_workbook(io.BytesIO(rendered.content))
    assert workbook.sheetnames == [DATA_SHEET, SOURCES_SHEET]


def test_headers_are_formatted_and_frozen() -> None:
    data, _ = sheets(write_xlsx(sample_table()).content)

    assert [cell.value for cell in data[1]] == [
        "Paper",
        "Source",
        "Pages",
        "Row status",
        "sample size",
        "dosing regimen",
    ]
    assert data["A1"].font.bold is True
    assert data["A1"].fill.fgColor.rgb == "0014181F"
    assert data.freeze_panes == "A2"
    assert data.auto_filter.ref == "A1:F1"


def test_values_and_metadata_land_in_the_right_cells() -> None:
    data, _ = sheets(write_xlsx(sample_table()).content)

    assert data["A2"].value == "A 6 Week Randomized Trial of Ziprasidone"
    assert data["B2"].hyperlink.target == "https://example.org/paper.pdf"
    assert data["C2"].value == 3
    assert data["D2"].value == "extracted"
    assert data["E2"].value == "73 patients"
    assert data["F2"].value == "40-160 mg/d for 6 weeks"


def test_each_cited_cell_links_to_its_sources_row() -> None:
    data, sources = sheets(write_xlsx(sample_table()).content)

    assert data["E2"].hyperlink.location == f"'{SOURCES_SHEET}'!A2"
    assert data["F2"].hyperlink.location == f"'{SOURCES_SHEET}'!A3"
    assert "73 patients were randomized" in data["E2"].hyperlink.tooltip

    assert [cell.value for cell in sources[1]][:7] == [
        "Ref",
        "Paper",
        "Column",
        "Value",
        "Page",
        "Quote match",
        "Quote",
    ]
    assert [sources[f"A{row}"].value for row in (2, 3)] == ["C1", "C2"]
    assert sources["C2"].value == "sample size"
    assert sources["E2"].value == 1
    assert sources["F2"].value == "exact wording"
    assert sources["G2"].value == "73 patients were randomized"
    assert sources["H2"].value == "p1-b4"
    assert sources["I2"].value == "58.1, 300.4, 296.1, 312.2"
    assert sources["J2"].hyperlink.target == "https://example.org/paper.pdf"


def test_uncited_values_are_marked_and_not_linked() -> None:
    data, sources = sheets(write_xlsx(sample_table()).content)

    assert data["E3"].value == "58 patients (no source found)"
    assert data["E3"].hyperlink is None
    assert data["E3"].font.italic is True
    assert data["F3"].value is None  # a not-found cell is written blank
    assert sources.max_row == 3  # only the two grounded cells


def test_fuzzy_matches_are_recorded_as_such() -> None:
    fuzzy = table(paper(cells={"sample_size": grounded("73", match=MatchQuality.FUZZY)}))

    data, sources = sheets(write_xlsx(fuzzy).content)

    assert sources["F2"].value == "close wording, not exact (fuzzy)"
    assert "close wording, not exact (fuzzy)" in data["E2"].hyperlink.tooltip


def test_formula_like_values_stay_text() -> None:
    injected = table(paper(cells={"sample_size": ungrounded('=HYPERLINK("http://evil")')}))

    data, _ = sheets(write_xlsx(injected).content)

    assert data["E2"].data_type == "s"
    assert data["E2"].value == '=HYPERLINK("http://evil") (no source found)'


def test_special_characters_survive_the_round_trip() -> None:
    awkward = 'IC₅₀ 3.4 µM ± 0.2, "high dose" — 中文, emoji 🧪'
    rendered = write_xlsx(table(paper(title=awkward, cells={"sample_size": ungrounded("n=1")})))

    data, _ = sheets(rendered.content)

    assert data["A2"].value == awkward


def test_control_characters_are_stripped_instead_of_crashing() -> None:
    row = table(paper(cells={"sample_size": grounded("73\x00 patients\x1f")}))

    data, _ = sheets(write_xlsx(row).content)

    assert data["E2"].value == "73 patients"


def test_values_longer_than_excels_cell_limit_are_truncated() -> None:
    row = table(paper(cells={"sample_size": ungrounded("x" * 40_000)}))

    data, _ = sheets(write_xlsx(row).content)

    assert data["E2"].value is not None
    assert len(data["E2"].value) == 32_767
    assert data["E2"].value.endswith("…")


def test_five_hundred_rows_render_quickly() -> None:
    big = table(
        *[
            paper(
                document_id=f"doc-{index}",
                title=f"Paper {index}",
                cells={
                    "sample_size": grounded(str(index)),
                    "dosing_regimen": grounded(f"{index} mg", page=2),
                },
            )
            for index in range(500)
        ]
    )

    started = time.monotonic()
    data, sources = sheets(write_xlsx(big).content)
    elapsed = time.monotonic() - started

    assert data.max_row == 501
    assert data["A501"].value == "Paper 499"
    assert sources.max_row == 1001
    assert sources["A1001"].value == "C1000"
    assert data["F501"].hyperlink.location == f"'{SOURCES_SHEET}'!A1001"
    assert elapsed < 30


def test_too_many_columns_is_rejected() -> None:
    wide = table(
        columns=[ExtractionField(key=f"f{i}", label=f"f{i}") for i in range(MAX_COLUMNS + 1)]
    )

    with pytest.raises(TableTooLargeError):
        write_xlsx(wide)


def test_citations_can_be_dropped_entirely() -> None:
    rendered = write_xlsx(sample_table(), ExportOptions(include_citations=False))

    workbook = load_workbook(io.BytesIO(rendered.content))

    assert workbook.sheetnames == [DATA_SHEET]
    assert workbook[DATA_SHEET]["E2"].hyperlink is None
