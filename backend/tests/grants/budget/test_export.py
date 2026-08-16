from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import load_workbook

from app.services.export import DATA_SHEET, SOURCES_SHEET, ExportFormat
from app.services.grants.budget import (
    BudgetCalculator,
    CostCategory,
    render,
    to_extraction_table,
)

from .conftest import cost, person, request


def budget(calculator: BudgetCalculator):  # type: ignore[no-untyped-def]
    return calculator.build(
        request(
            indirect_rate_percent=Decimal("40"),
            fee_percent=Decimal("7"),
            personnel=[person(), person(name="B. Ash", key_person=False)],
            costs=[
                cost(CostCategory.EQUIPMENT, "50000", "Plate reader"),
                cost(CostCategory.TRAVEL, "2400", "Conference travel"),
                cost(CostCategory.MATERIALS, "5000", "Reagents"),
            ],
        )
    )


def test_the_table_is_in_template_order_with_totals_in_place(
    calculator: BudgetCalculator,
) -> None:
    table = to_extraction_table(budget(calculator))
    codes = [row.title.split(".")[0] for row in table.rows]

    assert codes == sorted(codes)  # A..K, never out of order
    assert codes.index("G") < codes.index("H") < codes.index("I") < codes.index("J")
    assert [field.label for field in table.columns] == ["Line", "Basis", "Amount"]


def test_every_total_row_carries_the_calculated_amount(calculator: BudgetCalculator) -> None:
    costed = budget(calculator)
    amounts = {
        row.title.split(".")[0]: row.cells["amount"].value
        for row in to_extraction_table(costed).rows
    }
    assert amounts["G"] == f"{costed.total_direct:,}"
    assert amounts["I"] == f"{costed.total_direct_and_indirect:,}"
    assert amounts["K"] == f"{costed.total:,}"


def test_the_xlsx_comes_from_the_shared_export_service(calculator: BudgetCalculator) -> None:
    costed = budget(calculator)
    exported = render(costed)

    assert exported.filename == "grant-budget.xlsx"
    book = load_workbook(io.BytesIO(exported.content))
    # Reuse means the review table's sheet naming and header row come along unchanged, and the
    # citation sheet is absent because a budget line has no source document.
    assert book.sheetnames == [DATA_SHEET]
    assert SOURCES_SHEET not in book.sheetnames

    sheet = book[DATA_SHEET]
    assert [cell.value for cell in sheet[1]] == ["Section", "Line", "Basis", "Amount"]
    last = [cell.value for cell in sheet[sheet.max_row]]
    assert last[0].startswith("K.")
    assert last[3] == f"{costed.total:,}"


def test_the_csv_comes_from_the_shared_export_service(calculator: BudgetCalculator) -> None:
    exported = render(budget(calculator), ExportFormat.CSV)
    text = exported.content.decode("utf-8-sig")

    assert exported.filename == "grant-budget.csv"
    assert text.splitlines()[0] == "Section,Line,Basis,Amount"
    assert "K. Total Costs and Fee (I+J)" in text


def test_a_line_description_cannot_smuggle_a_formula_into_the_sheet(
    calculator: BudgetCalculator,
) -> None:
    costed = calculator.build(
        request(costs=[cost(CostCategory.MATERIALS, "100", "=cmd|'/c calc'!A0")])
    )
    sheet = load_workbook(io.BytesIO(render(costed).content))[DATA_SHEET]
    smuggled = [
        cell for row in sheet.iter_rows() for cell in row if str(cell.value).startswith("=cmd")
    ]
    assert smuggled and all(cell.data_type == "s" for cell in smuggled)

    csv_text = render(costed, ExportFormat.CSV).content.decode("utf-8-sig")
    assert "=cmd" in csv_text
    assert "\"'=cmd" in csv_text or "'=cmd" in csv_text
