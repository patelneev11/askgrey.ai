from __future__ import annotations

from io import BytesIO
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook

CREDENTIALS = {"email": "budget@askgrey.ai", "password": "obsidian-workspace-1"}


def auth_header(client: TestClient) -> dict[str, str]:
    tokens = client.post("/api/auth/register", json=CREDENTIALS).json()
    return {"Authorization": f"Bearer {tokens['access_token']}"}


def payload(**overrides: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "program": "SBIR",
        "phase": "phase_i",
        "period_months": 6,
        "organization": "Grey Therapeutics",
        "project_title": "Organoid screen",
        "personnel": [
            {
                "role": "Principal Investigator",
                "name": "A. Grey",
                "base_salary_annual": "120000",
                "effort_percent": "50",
                "months": "6",
                "fringe_rate_percent": "0",
            }
        ],
        "costs": [
            {"category": "materials", "description": "Assay plates", "unit_cost": "5000"},
        ],
        "indirect_rate_percent": "0",
        "fee_percent": "0",
    }
    base.update(overrides)
    return base


def test_budget_requires_authentication(client: TestClient) -> None:
    assert client.post("/api/grants/budget", json=payload()).status_code == 401


def test_budget_returns_sections_that_add_up_to_the_total(client: TestClient) -> None:
    response = client.post("/api/grants/budget", json=payload(), headers=auth_header(client))

    assert response.status_code == 200
    body = response.json()
    assert body["rules_version"]
    section_a = next(item for item in body["sections"] if item["code"] == "A")
    assert section_a["subtotal"] == "30000.00"
    # $120,000 x 50% x 6/12 plus $5,000 of materials, no indirect and no fee.
    assert body["total_direct"] == "35000.00"
    assert body["total"] == "35000.00"


def test_budget_reports_the_rules_it_applied(client: TestClient) -> None:
    response = client.post(
        "/api/grants/budget",
        json=payload(personnel=[dict(payload()["personnel"][0], base_salary_annual="900000")]),
        headers=auth_header(client),
    )

    assert response.status_code == 200
    assert "salary_cap" in {entry["rule_id"] for entry in response.json()["adjustments"]}


def test_budget_rejects_an_empty_request(client: TestClient) -> None:
    response = client.post(
        "/api/grants/budget",
        json=payload(personnel=[], costs=[]),
        headers=auth_header(client),
    )

    assert response.status_code == 422


def test_budget_rejects_an_impossible_effort(client: TestClient) -> None:
    response = client.post(
        "/api/grants/budget",
        json=payload(personnel=[dict(payload()["personnel"][0], effort_percent="140")]),
        headers=auth_header(client),
    )

    assert response.status_code == 422


def test_budget_rejects_more_personnel_lines_than_it_will_cost(client: TestClient) -> None:
    person = payload()["personnel"][0]
    response = client.post(
        "/api/grants/budget",
        json=payload(personnel=[person] * 51),
        headers=auth_header(client),
    )

    assert response.status_code == 422


def test_budget_rules_are_readable(client: TestClient) -> None:
    response = client.get("/api/grants/budget/rules", headers=auth_header(client))

    assert response.status_code == 200
    body = response.json()
    assert body["version"]
    assert float(body["salary_cap"]["annual_amount"]) > 0


def test_budget_exports_a_workbook_through_the_shared_exporter(client: TestClient) -> None:
    response = client.post(
        "/api/grants/budget/export",
        params={"format": "xlsx"},
        json=payload(),
        headers=auth_header(client),
    )

    assert response.status_code == 200
    assert "grant-budget" in response.headers["content-disposition"]
    sheet = load_workbook(BytesIO(response.content)).worksheets[0]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    assert any(row and row[0] == "K. Total Costs and Fee (I+J)" for row in rows)


def test_budget_export_rejects_an_empty_request(client: TestClient) -> None:
    response = client.post(
        "/api/grants/budget/export",
        json=payload(personnel=[], costs=[]),
        headers=auth_header(client),
    )

    assert response.status_code == 422


def test_budget_export_requires_authentication(client: TestClient) -> None:
    assert client.post("/api/grants/budget/export", json=payload()).status_code == 401
